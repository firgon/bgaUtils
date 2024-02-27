from openpyxl import Workbook, worksheet, workbook, load_workbook
import pandas as pd
from .Data import Data
from utils.Utils import promptChoice


def run():
    # file = Utils.choose_file("Choississez le fichier de données", ("Xls files", "*.xlsx"))
    file = "C:/Users/Manu/Downloads/all_cards.xlsx"

    xls = pd.ExcelFile(file)

    excel = pd.read_excel(file, sheet_name="20_Ereignisse_Events")

    ID = '#'

    myData = Data([ID, 'Name EN', 'Text EN'])

    sheet = pd.DataFrame(excel, columns=myData.fieldNames)

    for index, row in sheet.iterrows():
        data = dict()
        if pd.isna(row[ID]):
            print("cette cellule est vide", index)
            continue

        for d in myData.fieldNames:
            if d == ID:
                data[d] = cleanId(row[d])
            else:
                value = row[d]
                if index+1 in sheet[ID].keys() and pd.isna(sheet[ID][index+1]) and pd.notna(sheet[ID][index+1]):
                    value += " " + sheet[d][index+1]
                data[d] = value
        myData.add(data)




def cleanId(id: str) -> int:
    return int(id[1:])


def selectSheet(name: str, wb: workbook) -> worksheet:
    for ws in wb.sheetnames:
        if name in ws:
            return ws

    print("Je ne trouve pas la bonne feuille : ", name)
    print(wb.sheetnames)
    exit()


if __name__ == '__main__':
    run()
